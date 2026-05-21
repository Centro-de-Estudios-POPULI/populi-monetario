"""
Scraper: Financiamiento del BCB al Sector Público (Monetización de Deuda)
Fuente: BCB — https://www.bcb.gob.bo/?q=content/sector-monetario
Genera: data/monetizacion_deuda.json
"""

import json, os, sys, datetime, time
from pathlib import Path
import openpyxl
import requests

URL = "https://www.bcb.gob.bo/webdocs/sector_monetario/Cr%C3%A9ditos%20y%20Dep%C3%B3sitos/BCB/3.%20Financiamiento%20neto%20al%20Sector%20P%C3%BAblico.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = Path(__file__).resolve().parent / "bcb_raw" / "financiamiento_spnf.xlsx"


def download():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"Descargando {URL}")
    for attempt in range(3):
        try:
            r = requests.get(URL, timeout=90)
            r.raise_for_status()
            XLSX_PATH.write_bytes(r.content)
            print(f"  -> {XLSX_PATH} ({len(r.content)//1024} KB)")
            return
        except Exception as e:
            delay = [10, 30, 60][attempt]
            if attempt < 2:
                print(f"  Intento {attempt+1}/3: {e} — reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"  Error tras 3 intentos: {e}")
                sys.exit(1)


def to_mm(v):
    return round(float(v) / 1000, 2) if v and isinstance(v, (int, float)) else 0


def parse():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb[wb.sheetnames[0]]

    series = []
    for r in range(9, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if date_val is None:
            continue
        if isinstance(date_val, datetime.datetime):
            date_str = date_val.strftime("%Y-%m")
        elif isinstance(date_val, str) and len(date_val) >= 7:
            date_str = date_val[:7]
        else:
            continue

        gc_bruto = ws.cell(r, 2).value
        if gc_bruto is None or not isinstance(gc_bruto, (int, float)):
            continue

        series.append({
            "date": date_str,
            "gc_bruto": to_mm(gc_bruto),
            "gc_depositos": to_mm(ws.cell(r, 3).value),
            "gc_neto": to_mm(ws.cell(r, 4).value),
            "ss_neto": to_mm(ws.cell(r, 7).value),
            "gl_neto": to_mm(ws.cell(r, 10).value),
            "ep_neto": to_mm(ws.cell(r, 13).value),
            "total_bruto": to_mm(ws.cell(r, 14).value),
            "total_depositos": to_mm(ws.cell(r, 15).value),
            "total_neto": to_mm(ws.cell(r, 16).value),
        })

    last = series[-1]
    prev_12 = series[-13] if len(series) >= 13 else series[0]

    def var_abs(key):
        return round(last[key] - prev_12[key], 2) if last[key] and prev_12[key] else None

    metadata = {
        "titulo": "Monetización de la Deuda Pública",
        "subtitulo": "Financiamiento del BCB al Sector Público No Financiero",
        "fuente": "Banco Central de Bolivia (BCB)",
        "unidad": "Millones de Bs",
        "frecuencia": "Mensual",
        "ultimo_dato": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "credito_neto_total_mm": last["total_neto"],
        "credito_neto_gc_mm": last["gc_neto"],
        "var_12m_neto_abs": var_abs("total_neto"),
    }

    return {"metadata": metadata, "series": series}


def main():
    if "--no-download" not in sys.argv:
        download()
    data = parse()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "monetizacion_deuda.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    m = data["metadata"]
    print(f"OK: {m['observaciones']} obs, {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"   Crédito Neto SPNF: {m['credito_neto_total_mm']:,.0f} MM Bs | GC Neto: {m['credito_neto_gc_mm']:,.0f} MM Bs")


if __name__ == "__main__":
    main()
