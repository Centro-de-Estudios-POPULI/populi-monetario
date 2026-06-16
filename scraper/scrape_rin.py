"""
Scraper: Reservas Internacionales (RIN/RIB)
Fuente: BCB — Estadísticas Semanales
Genera: data/reservas_rin.json

El BCB publica un Excel por semana en:
  https://www.bcb.gob.bo/webdocs/05_estadisticassemanales/Semanal%20{SEMANA}_{AÑO}.xlsx
El scraper busca automáticamente la semana más reciente disponible.
"""

import json, os, sys, datetime, time, unicodedata
from pathlib import Path
import openpyxl
import requests

BASE_URL = "https://www.bcb.gob.bo/webdocs/05_estadisticassemanales/Semanal%20{week}_{year}.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = Path(__file__).resolve().parent / "bcb_raw" / "semanal.xlsx"

# El BCB inserta/mueve filas en el Excel semanal sin previo aviso (p.ej. en
# 2026 agregó "Recursos de Alta Liquidez" y "Oro convertible en divisas",
# desplazando el bloque de reservas +2 filas). Por eso NO usamos números de
# fila fijos: localizamos cada renglón por su etiqueta en find_row_map().
def _norm(s) -> str:
    """Minúsculas sin acentos ni espacios extra, para comparar etiquetas."""
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split()).lower()


def _label(ws, r) -> str:
    """Primera celda de texto no vacía en las columnas 1–4 de la fila r."""
    for cc in range(1, 5):
        v = ws.cell(r, cc).value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def find_row_map(ws) -> dict:
    """
    Ubica las filas de la sección de reservas por etiqueta (no por posición).
    Ancla en 'Reservas internacionales brutas' .. 'Reservas internacionales
    netas' y busca dentro de ese bloque las partidas exactas. Lanza error si
    falta alguna (así el workflow falla en vez de publicar datos desalineados).
    """
    rib = rin = None
    for r in range(1, ws.max_row + 1):
        lab = _norm(_label(ws, r))
        if rib is None and "reservas internacionales brutas" in lab:
            rib = r
        elif "reservas internacionales netas" in lab:
            rin = r
            break

    if not rib or not rin or rin <= rib:
        raise ValueError(
            f"No se ubicaron las filas RIB/RIN en el Excel del BCB "
            f"(rib={rib}, rin={rin}); el formato pudo cambiar."
        )

    rows = {"rib": rib, "rin": rin}
    for r in range(rib + 1, rin):           # solo dentro del bloque de reservas
        lab = _norm(_label(ws, r))
        if "deg" not in rows and lab == "deg":
            rows["deg"] = r
        elif "oro" not in rows and lab == "oro":            # exacto: evita "d/c Oro…", "Oro convertible…"
            rows["oro"] = r
        elif "divisas" not in rows and lab == "divisas":    # exacto: evita "Oro convertible en divisas"
            rows["divisas"] = r
        elif "fmi" not in rows and "posicion con el fmi" in lab:
            rows["fmi"] = r

    faltan = [k for k in ("deg", "oro", "divisas", "fmi") if k not in rows]
    if faltan:
        raise ValueError(
            f"No se ubicaron filas {faltan} en el bloque de reservas del BCB; "
            f"las etiquetas pudieron cambiar."
        )
    return rows

# El WAF del BCB puede rechazar (403) el User-Agent por defecto de python-requests.
# Emulamos un navegador para reducir los bloqueos.
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
              "application/vnd.ms-excel,*/*",
    "Referer": "https://www.bcb.gob.bo/",
}


def find_latest_url():
    now = datetime.date.today()
    year = now.year
    week = now.isocalendar()[1]

    suffixes = ["", "_0", "%20(1)"]

    for w in range(week, max(week - 6, 0), -1):
        for suffix in suffixes:
            url = BASE_URL.format(week=w, year=year) .replace(".xlsx", f"{suffix}.xlsx")
            try:
                r = requests.head(url, headers=HEADERS, timeout=15, allow_redirects=True)
                if r.status_code == 200:
                    print(f"  Encontrado: Semana {w}/{year}")
                    return url
            except:
                continue

    if now.month == 1 and week <= 2:
        for w in range(53, 48, -1):
            url = BASE_URL.format(week=w, year=year-1)
            try:
                r = requests.head(url, headers=HEADERS, timeout=15, allow_redirects=True)
                if r.status_code == 200:
                    return url
            except:
                continue

    return None


def download():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    url = find_latest_url()
    if not url:
        print("  No se encontró ningún Excel semanal reciente del BCB")
        sys.exit(1)

    print(f"Descargando {url}")
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=90)
            r.raise_for_status()
            XLSX_PATH.write_bytes(r.content)
            print(f"  -> {XLSX_PATH} ({len(r.content)//1024} KB)")
            return
        except Exception as e:
            delay = [10, 30, 60][attempt]
            if attempt < 2:
                print(f"  Intento {attempt+1}/3: {e} — reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"  Error tras 3 intentos: {e}")
                sys.exit(1)


def parse():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb[wb.sheetnames[0]]

    row_map = find_row_map(ws)
    print(f"  Filas detectadas: {row_map}")

    series = []
    last_month_date = None

    for c in range(5, ws.max_column + 1):
        date_val = ws.cell(3, c).value
        if date_val is None:
            continue

        if isinstance(date_val, datetime.datetime):
            date_str = date_val.strftime("%Y-%m")
            last_month_date = date_val
        elif isinstance(date_val, str) and "semana" in date_val.lower():
            if last_month_date is None:
                continue
            next_m = last_month_date.month + 1
            next_y = last_month_date.year
            if next_m > 12:
                next_m = 1
                next_y += 1
            week_num = int("".join(filter(str.isdigit, date_val)) or "1")
            date_str = f"{next_y}-{next_m:02d}-S{week_num}"
        else:
            continue

        row_data = {}
        all_none = True
        for key, row in row_map.items():
            v = ws.cell(row, c).value
            if v is not None and isinstance(v, (int, float)):
                row_data[key] = round(float(v), 2)
                all_none = False
            else:
                row_data[key] = None

        if all_none:
            continue

        entry = {"date": date_str}
        entry.update(row_data)
        series.append(entry)

    monthly = [s for s in series if "-S" not in s["date"]]
    weekly = [s for s in series if "-S" in s["date"]]

    last = monthly[-1] if monthly else series[-1]
    prev_12 = monthly[-13] if len(monthly) >= 13 else monthly[0]

    # Validación de coherencia: RIB ≈ Oro + Divisas + DEG + Posición FMI.
    # Si las filas quedaran mal mapeadas (el BCB volvió a mover el Excel), la
    # identidad se rompe y abortamos en vez de publicar cifras desalineadas.
    comp = [last.get(k) for k in ("oro", "divisas", "deg", "fmi")]
    if last.get("rib") and all(v is not None for v in comp):
        suma = sum(comp)
        rib = last["rib"]
        if abs(suma - rib) > max(5.0, rib * 0.02):   # tolerancia: 2% o 5 MM USD
            raise ValueError(
                f"Identidad de reservas rota en {last['date']}: "
                f"Oro+Divisas+DEG+FMI={suma:.1f} ≠ RIB={rib:.1f}. "
                f"Revisar el mapeo de filas del Excel del BCB."
            )

    def safe_var(cur, prev):
        if cur and prev and prev > 0:
            return round((cur / prev - 1) * 100, 1)
        return None

    latest_point = weekly[-1] if weekly else last

    metadata = {
        "titulo": "Reservas Internacionales",
        "subtitulo": "Reservas Internacionales Netas y Brutas del BCB (en millones de USD)",
        "fuente": "Banco Central de Bolivia (BCB) — Estadísticas Semanales",
        "unidad": "Millones de USD",
        "frecuencia": "Mensual + Semanal",
        "ultimo_dato": latest_point["date"],
        "ultimo_mensual": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "observaciones_mensuales": len(monthly),
        "ultimo_rin": latest_point.get("rin"),
        "ultimo_rib": latest_point.get("rib"),
        "var_12m_rin_pct": safe_var(last.get("rin"), prev_12.get("rin")),
    }

    return {"metadata": metadata, "series": series}


def main():
    if "--no-download" not in sys.argv:
        download()
    data = parse()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "reservas_rin.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    m = data["metadata"]
    print(f"OK: {m['observaciones']} obs ({m['observaciones_mensuales']} mensuales), {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"   RIN: ${m['ultimo_rin']:,.1f} MM USD | Var 12m: {m['var_12m_rin_pct']}%")


if __name__ == "__main__":
    main()
