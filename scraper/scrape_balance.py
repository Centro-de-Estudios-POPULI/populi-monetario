"""
Scraper: Balance del BCB (Activo y Pasivo)
Fuente: BCB — https://www.bcb.gob.bo/?q=content/sector-monetario
Genera: data/activo_bcb.json, data/pasivo_bcb.json
"""

import json, os, sys, datetime, time
from pathlib import Path
import openpyxl
import requests

URL = "https://www.bcb.gob.bo/webdocs/sector_monetario/Balances%20Consolidados/4.%20Banco%20Central.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = Path(__file__).resolve().parent / "bcb_raw" / "balance_bcb.xlsx"


def download():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"Descargando {URL}")
    for attempt in range(3):
        try:
            r = requests.get(URL, timeout=90)
            r.raise_for_status()
            XLSX_PATH.write_bytes(r.content)
            print(f"  -> {XLSX_PATH} ({len(r.content)//1024} KB)")
            return
        except Exception as e:
            delay = [10, 30, 60][attempt]
            if attempt < 2:
                print(f"  Intento {attempt+1}/3: {e} — reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"  Error tras 3 intentos: {e}")
                sys.exit(1)


def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m")
    if isinstance(val, (int, float)):
        return f"{int(val)}-12"
    return None


def to_mm(v):
    return round(float(v) / 1000, 2) if v and isinstance(v, (int, float)) else 0


def parse_activo():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb["Activo"]

    series = []
    for r in range(11, ws.max_row + 1):
        date_str = parse_date(ws.cell(r, 1).value)
        if not date_str:
            continue

        rib = ws.cell(r, 2).value
        if rib is None or not isinstance(rib, (int, float)):
            continue

        series.append({
            "date": date_str,
            "rib": to_mm(rib),
            "oro": to_mm(ws.cell(r, 3).value),
            "divisas": to_mm(ws.cell(r, 4).value),
            "otros_rib": to_mm(ws.cell(r, 5).value),
            "aportes_oi": to_mm(ws.cell(r, 6).value),
            "otros_ext": to_mm(ws.cell(r, 7).value),
            "credito_sp": to_mm(ws.cell(r, 8).value),
            "credito_gc": to_mm(ws.cell(r, 9).value),
            "credito_ss": to_mm(ws.cell(r, 10).value),
            "credito_gl": to_mm(ws.cell(r, 11).value),
            "credito_ep": to_mm(ws.cell(r, 12).value),
            "credito_sf": to_mm(ws.cell(r, 13).value),
            "otras_ctas": to_mm(ws.cell(r, 16).value),
            "total_activo": to_mm(ws.cell(r, 17).value),
        })

    last = series[-1]
    prev_12 = series[-13] if len(series) >= 13 else series[0]

    metadata = {
        "titulo": "Activo del Banco Central de Bolivia",
        "subtitulo": "Composición del activo: reservas, crédito al sector público y financiero",
        "fuente": "Banco Central de Bolivia (BCB)",
        "unidad": "Millones de Bs",
        "frecuencia": "Mensual",
        "ultimo_dato": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "total_activo_mm": last["total_activo"],
        "rib_mm": last["rib"],
        "credito_sp_mm": last["credito_sp"],
        "credito_sf_mm": last["credito_sf"],
    }

    return {"metadata": metadata, "series": series}


def parse_pasivo():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb["Pasivo"]

    series = []
    for r in range(12, ws.max_row + 1):
        date_str = parse_date(ws.cell(r, 1).value)
        if not date_str:
            continue

        emision = ws.cell(r, 2).value
        if emision is None or not isinstance(emision, (int, float)):
            continue

        series.append({
            "date": date_str,
            "emision": to_mm(emision),
            "dep_bancarios": to_mm(ws.cell(r, 3).value),
            "oblig_ext_cp": to_mm(ws.cell(r, 6).value),
            "dep_sp_total": to_mm(ws.cell(r, 9).value),
            "dep_gc": to_mm(ws.cell(r, 10).value),
            "dep_ss": to_mm(ws.cell(r, 16).value),
            "dep_gl": to_mm(ws.cell(r, 19).value),
            "dep_ep": to_mm(ws.cell(r, 22).value),
            "dep_oi": to_mm(ws.cell(r, 25).value),
            "oblig_ext_mlp": to_mm(ws.cell(r, 26).value),
            "otras_ctas": to_mm(ws.cell(r, 27).value),
            "capital": to_mm(ws.cell(r, 31).value),
        })

    last = series[-1]

    metadata = {
        "titulo": "Pasivo del Banco Central de Bolivia",
        "subtitulo": "Composición del pasivo: emisión, depósitos, obligaciones y capital",
        "fuente": "Banco Central de Bolivia (BCB)",
        "unidad": "Millones de Bs",
        "frecuencia": "Mensual",
        "ultimo_dato": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "emision_mm": last["emision"],
        "dep_sp_mm": last["dep_sp_total"],
        "dep_bancarios_mm": last["dep_bancarios"],
    }

    return {"metadata": metadata, "series": series}


def main():
    if "--no-download" not in sys.argv:
        download()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    activo = parse_activo()
    out1 = OUT_DIR / "activo_bcb.json"
    with open(out1, "w", encoding="utf-8") as f:
        json.dump(activo, f, ensure_ascii=False, indent=2)
    m = activo["metadata"]
    print(f"ACTIVO: {m['observaciones']} obs, {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"  Total Activo: {m['total_activo_mm']:,.0f} MM | RIB: {m['rib_mm']:,.0f} MM | Créd SP: {m['credito_sp_mm']:,.0f} MM")

    pasivo = parse_pasivo()
    out2 = OUT_DIR / "pasivo_bcb.json"
    with open(out2, "w", encoding="utf-8") as f:
        json.dump(pasivo, f, ensure_ascii=False, indent=2)
    m = pasivo["metadata"]
    print(f"PASIVO: {m['observaciones']} obs, {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"  Emisión: {m['emision_mm']:,.0f} MM | Dep SP: {m['dep_sp_mm']:,.0f} MM | Dep Banc: {m['dep_bancarios_mm']:,.0f} MM")


if __name__ == "__main__":
    main()
