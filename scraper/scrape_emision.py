"""
Scraper: Emisión Monetaria (Destino del medio circulante y de la liquidez total)
Fuente: BCB — https://www.bcb.gob.bo/?q=content/sector-monetario
Genera: data/emision_monetaria.json
"""

import json, os, sys, datetime, time
from pathlib import Path
import openpyxl
import requests

URL = "https://www.bcb.gob.bo/webdocs/sector_monetario/Indicadores%20Monetarios/6.%20Destino%20del%20medio%20circulante%20y%20la%20liquidez%20total.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = Path(__file__).resolve().parent / "bcb_raw" / "destino_circulante.xlsx"


def download():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"Descargando {URL}")
    for attempt in range(3):
        try:
            r = requests.get(URL, timeout=90)
            r.raise_for_status()
            XLSX_PATH.write_bytes(r.content)
            print(f"  -> {XLSX_PATH} ({len(r.content)//1024} KB)")
            return
        except Exception as e:
            delay = [10, 30, 60][attempt]
            if attempt < 2:
                print(f"  Intento {attempt+1}/3: {e} — reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"  Error tras 3 intentos: {e}")
                sys.exit(1)


def parse():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb[wb.sheetnames[0]]

    series = []
    for r in range(12, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if date_val is None:
            continue
        if isinstance(date_val, datetime.datetime):
            date_str = date_val.strftime("%Y-%m")
        elif isinstance(date_val, str) and len(date_val) >= 7:
            date_str = date_val[:7]
        else:
            continue

        emision = ws.cell(r, 2).value
        caja_sf = ws.cell(r, 3).value
        billetes = ws.cell(r, 4).value

        if emision is None or not isinstance(emision, (int, float)):
            continue

        series.append({
            "date": date_str,
            "emision": round(float(emision) / 1000, 2),
            "caja_sf": round(float(caja_sf) / 1000, 2) if caja_sf else 0,
            "billetes": round(float(billetes) / 1000, 2) if billetes else 0,
        })

    last = series[-1]
    prev_12 = series[-13] if len(series) >= 13 else series[0]

    metadata = {
        "titulo": "Emisión Monetaria",
        "subtitulo": "Destino del medio circulante: billetes en poder del público y caja del sistema financiero",
        "fuente": "Banco Central de Bolivia (BCB)",
        "unidad": "Millones de Bs",
        "frecuencia": "Mensual",
        "ultimo_dato": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "ultima_emision_mm": last["emision"],
        "var_12m_pct": round((last["emision"] / prev_12["emision"] - 1) * 100, 1) if prev_12["emision"] else None,
    }

    return {"metadata": metadata, "series": series}


def main():
    if "--no-download" not in sys.argv:
        download()
    data = parse()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "emision_monetaria.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    m = data["metadata"]
    print(f"OK: {m['observaciones']} obs, {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"   Última emisión: {m['ultima_emision_mm']:,.0f} MM Bs | Var 12m: {m['var_12m_pct']}%")


if __name__ == "__main__":
    main()
