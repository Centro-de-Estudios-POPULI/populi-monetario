"""
Scraper: Agregados Monetarios (M1, M2, M3, M4)
Fuente: BCB — https://www.bcb.gob.bo/?q=content/sector-monetario
Genera: data/agregados_monetarios.json
"""

import json, os, sys, datetime, time
from pathlib import Path
import openpyxl
import requests

URL = "https://www.bcb.gob.bo/webdocs/sector_monetario/Indicadores%20Monetarios/7.%20Agregados%20Monetarios.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = Path(__file__).resolve().parent / "bcb_raw" / "agregados_monetarios.xlsx"

# El WAF del BCB puede rechazar (403) el User-Agent por defecto de python-requests.
# Emulamos un navegador para reducir los bloqueos.
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
              "application/vnd.ms-excel,*/*",
    "Referer": "https://www.bcb.gob.bo/",
}


def download():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"Descargando {URL}")
    for attempt in range(3):
        try:
            r = requests.get(URL, headers=HEADERS, timeout=90)
            r.raise_for_status()
            XLSX_PATH.write_bytes(r.content)
            print(f"  -> {XLSX_PATH} ({len(r.content)//1024} KB)")
            return
        except Exception as e:
            delay = [10, 30, 60][attempt]
            if attempt < 2:
                print(f"  Intento {attempt+1}/3: {e} — reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"  Error tras 3 intentos: {e}")
                sys.exit(1)


def parse():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb[wb.sheetnames[0]]

    series = []
    for r in range(12, ws.max_row + 1):
        date_val = ws.cell(r, 2).value
        if date_val is None:
            continue
        if isinstance(date_val, datetime.datetime):
            date_str = date_val.strftime("%Y-%m")
        elif isinstance(date_val, str) and len(date_val) >= 7:
            date_str = date_val[:7]
        else:
            continue

        billetes = ws.cell(r, 3).value
        m1 = ws.cell(r, 4).value
        m2 = ws.cell(r, 6).value
        m3 = ws.cell(r, 8).value
        m4 = ws.cell(r, 10).value

        if m1 is None or not isinstance(m1, (int, float)):
            continue

        def to_mm(v):
            return round(float(v) / 1000, 2) if v and isinstance(v, (int, float)) else 0

        series.append({
            "date": date_str,
            "billetes": to_mm(billetes),
            "m1": to_mm(m1),
            "m2": to_mm(m2),
            "m3": to_mm(m3),
            "m4": to_mm(m4),
        })

    last = series[-1]
    prev_12 = series[-13] if len(series) >= 13 else series[0]

    def var12(key):
        if prev_12[key] and prev_12[key] > 0:
            return round((last[key] / prev_12[key] - 1) * 100, 1)
        return None

    metadata = {
        "titulo": "Agregados Monetarios",
        "subtitulo": "M1, M2, M3, M4 — Saldos a fin de período",
        "fuente": "Banco Central de Bolivia (BCB)",
        "unidad": "Millones de Bs",
        "frecuencia": "Mensual",
        "ultimo_dato": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "ultimo_m4_mm": last["m4"],
        "var_12m_m1_pct": var12("m1"),
        "var_12m_m2_pct": var12("m2"),
        "var_12m_m3_pct": var12("m3"),
        "var_12m_m4_pct": var12("m4"),
    }

    return {"metadata": metadata, "series": series}


def main():
    if "--no-download" not in sys.argv:
        download()
    data = parse()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "agregados_monetarios.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    m = data["metadata"]
    print(f"OK: {m['observaciones']} obs, {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"   M4: {m['ultimo_m4_mm']:,.0f} MM Bs | Var 12m M4: {m['var_12m_m4_pct']}%")


if __name__ == "__main__":
    main()
