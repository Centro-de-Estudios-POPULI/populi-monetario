"""
Scraper: Base Monetaria — Determinantes
Fuente: BCB — "2. Base Monetaria, Determinantes y Componentes.xlsx"
        https://www.bcb.gob.bo/?q=content/sector-monetario
Genera: data/determinantes_base.json

Este es el cuadro OFICIAL del BCB de los determinantes de la base monetaria.
Su identidad (lado de los determinantes):

    Base Monetaria = Reservas Internacionales Netas
                   + Crédito Neto al Sector Público
                   + Crédito a Bancos
                   - OMA                          (operaciones de mercado abierto; se RESTAN)
                   + Otras Cuentas (Neto)

y por el lado de los componentes: Base Monetaria = Billetes y Monedas + Reservas Bancarias.

El cuadro está TRANSPUESTO: las fechas corren por columnas (desde 2001-12) y los
rubros por fila. Los rubros se ubican por ETIQUETA dentro de la banda
DETERMINANTES (nunca por número de fila fijo), porque el BCB inserta/desplaza
filas entre versiones. Definición de base monetaria distinta (y más amplia) que
la "Emisión": Base = Billetes + Reservas Bancarias (incluye encaje en el BCB),
mientras que Emisión = Billetes + Caja.
"""

import json, re, sys, datetime, time, unicodedata
from pathlib import Path
import openpyxl
import requests

URL = "https://www.bcb.gob.bo/webdocs/sector_monetario/Indicadores%20Monetarios/2.%20Base%20Monetaria,%20Determinantes%20y%20Componentes.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = Path(__file__).resolve().parent / "bcb_raw" / "base_determinantes.xlsx"

# El WAF del BCB puede rechazar (403) el User-Agent por defecto de python-requests.
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
              "application/vnd.ms-excel,*/*",
    "Referer": "https://www.bcb.gob.bo/",
}


def download():
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"Descargando {URL}")
    for attempt in range(3):
        try:
            r = requests.get(URL, headers=HEADERS, timeout=90)
            r.raise_for_status()
            XLSX_PATH.write_bytes(r.content)
            print(f"  -> {XLSX_PATH} ({len(r.content)//1024} KB)")
            return
        except Exception as e:
            delay = [10, 30, 60][attempt]
            if attempt < 2:
                print(f"  Intento {attempt+1}/3: {e} — reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"  Error tras 3 intentos: {e}")
                sys.exit(1)


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return " ".join(s.lower().split())


def find_rows(ws):
    """Ubica las filas de cada rubro por etiqueta dentro de la banda DETERMINANTES."""
    labels = {}
    r_det = r_comp = None
    for r in range(1, ws.max_row + 1):
        t = norm(ws.cell(r, 1).value)
        if t == "determinantes":
            r_det = r
        elif t == "componentes":
            r_comp = r
            break
    if r_det is None or r_comp is None:
        sys.exit("ERROR: no se encontraron las bandas DETERMINANTES/COMPONENTES")

    # Mapa etiqueta-normalizada -> clave de salida (rubros de los determinantes)
    det_map = {
        "reservas internacionales netas": "rin",
        "credito neto al sector publico": "cred_sp",
        "credito a bancos": "cred_bancos",
        "otras cuentas (neto)": "otras",
    }
    for r in range(r_det + 1, r_comp):
        t = norm(ws.cell(r, 1).value)
        if t in det_map:
            labels[det_map[t]] = r
        elif t.startswith("oma"):           # "OMA 1" (con marca de nota al pie)
            labels.setdefault("oma", r)

    # Componentes (para verificar la identidad)
    for r in range(r_comp + 1, ws.max_row + 1):
        t = norm(ws.cell(r, 1).value)
        if t == "billetes y monedas":
            labels["billetes"] = r
        elif t == "reservas bancarias":
            labels["reservas_bancarias"] = r
            break

    req = ["rin", "cred_sp", "cred_bancos", "oma", "otras", "billetes", "reservas_bancarias"]
    missing = [k for k in req if k not in labels]
    if missing:
        sys.exit(f"ERROR: faltan rubros por etiqueta: {missing}")
    return labels


def to_mm(v):
    return round(float(v) / 1000, 1) if isinstance(v, (int, float)) else None


MESES = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
         "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}


def col_date(v):
    """Fecha de una columna como 'AAAA-MM'. El BCB mezcla dos formatos en la misma
    fila: celdas de tipo fecha para la historia y ETIQUETAS DE TEXTO ('abr26',
    'may26', 'jun26') para los meses más recientes. Leer solo las de tipo fecha
    dejaba fuera los últimos meses publicados."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m")
    if isinstance(v, str):
        s = norm(v).replace("-", "").replace(" ", "").replace(".", "").replace("/", "")
        m = re.fullmatch(r"([a-z]{3,4})(\d{2,4})", s)
        if m and m.group(1)[:3] in MESES:
            y = int(m.group(2))
            y = y + 2000 if y < 100 else y
            return f"{y}-{MESES[m.group(1)[:3]]:02d}"
    return None


def find_date_row(ws):
    """Fila de las fechas (cabecera del cuadro transpuesto). NO se fija por número:
    el BCB la mueve al insertar/quitar filas de título (jul-2026 pasó de la 6 a la
    5 y el parser devolvía una serie vacía). Se elige la fila con más celdas
    interpretables como fecha entre las primeras 15."""
    mejor, n_mejor = None, 0
    for r in range(1, 16):
        n = sum(1 for c in range(3, ws.max_column + 1)
                if col_date(ws.cell(r, c).value))
        if n > n_mejor:
            mejor, n_mejor = r, n
    if not mejor or n_mejor < 12:
        sys.exit("ERROR: no se encontró la fila de fechas del cuadro")
    return mejor


def parse():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb["2"] if "2" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = find_rows(ws)
    r_fechas = find_date_row(ws)

    series = []
    max_dif = 0.0
    for c in range(3, ws.max_column + 1):
        date_str = col_date(ws.cell(r_fechas, c).value)
        if not date_str:
            continue

        def g(key):
            return to_mm(ws.cell(rows[key], c).value)

        rin, cred_sp = g("rin"), g("cred_sp")
        cred_bancos, oma, otras = g("cred_bancos"), g("oma"), g("otras")
        billetes, res_banc = g("billetes"), g("reservas_bancarias")
        if rin is None or billetes is None:
            continue

        oma = oma or 0
        base_det = rin + cred_sp + cred_bancos - oma + otras
        base_comp = billetes + res_banc
        max_dif = max(max_dif, abs(base_det - base_comp))

        series.append({
            "date": date_str,
            "base": round(base_det, 1),
            "rin": rin,
            "cred_sp": cred_sp,
            "cred_bancos": cred_bancos,
            "oma": oma,                      # publicado en positivo; es contractivo (se resta)
            "otras": otras,
            "billetes": billetes,
            "reservas_bancarias": res_banc,
        })

    if max_dif > 1.0:
        print(f"  ADVERTENCIA: determinantes vs componentes no cuadran (dif. máx. {max_dif:.2f} MM Bs)")

    last = series[-1]
    prev_12 = series[-13] if len(series) >= 13 else series[0]

    def var12(key):
        if prev_12[key] and abs(prev_12[key]) > 0:
            return round((last[key] / prev_12[key] - 1) * 100, 1)
        return None

    metadata = {
        "titulo": "Determinantes de la Base Monetaria",
        "subtitulo": "Fuentes de creación de dinero primario según el cuadro oficial del BCB: reservas, crédito al sector público, crédito a bancos y OMA",
        "fuente": "Banco Central de Bolivia (BCB)",
        "unidad": "Millones de Bs",
        "frecuencia": "Mensual",
        "ultimo_dato": last["date"],
        "primer_dato": series[0]["date"],
        "observaciones": len(series),
        "base_mm": last["base"],
        "rin_mm": last["rin"],
        "cred_sp_mm": last["cred_sp"],
        "oma_mm": last["oma"],
        "share_rin_pct": round(last["rin"] / last["base"] * 100, 1) if last["base"] else None,
        "share_cred_sp_pct": round(last["cred_sp"] / last["base"] * 100, 1) if last["base"] else None,
        "var_12m_base_pct": var12("base"),
        "var_12m_rin_pct": var12("rin"),
        "var_12m_cred_sp_pct": var12("cred_sp"),
        "verif_dif_max_mm": round(max_dif, 3),
    }

    return {"metadata": metadata, "series": series}


def main():
    if "--no-download" not in sys.argv:
        download()
    data = parse()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "determinantes_base.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    m = data["metadata"]
    print(f"OK: {m['observaciones']} obs, {m['primer_dato']} a {m['ultimo_dato']}")
    print(f"   Base Monetaria: {m['base_mm']:,.0f} MM | RIN: {m['rin_mm']:,.0f} ({m['share_rin_pct']}%) "
          f"| Créd. SP: {m['cred_sp_mm']:,.0f} ({m['share_cred_sp_pct']}%) | OMA: -{m['oma_mm']:,.0f}")
    print(f"   Identidad determinantes=componentes: dif. máx. {m['verif_dif_max_mm']} MM Bs")


if __name__ == "__main__":
    main()
